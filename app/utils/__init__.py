from collections import defaultdict
from decimal import Decimal
from openpyxl import load_workbook
from PIL.Image import open as open_image
from pyglet.canvas import get_display
from pyglet.graphics import Batch
from pyglet.window import Window
from typing import Any, Iterable

from library import (
    Facility,
    Grid
)

def center(window: Window):
    screen = get_display().get_screens()[0]
    x = screen.width // 2 - window.width // 2
    window.set_location(x, 50)

def interpret(value):
    if value is None:
        return None
    if any([
        isinstance(value, float),
        isinstance(value, int),
        isinstance(value, str) and (value.isdigit() or value.isdecimal())
    ]):
        return Decimal(value)
    return str(value)

def parse_xlsx(path: str, *sheetnames):
    workbook = load_workbook(path)
    worksheets: list[list[list]] = []
    for name in sheetnames:
        worksheet = workbook[name]
        data: list[list] = []
        for row in worksheet.rows:
            data.append(list(interpret(cell.value) for cell in row))
        worksheets.append(data)
    workbook.close()
    return worksheets

def initialize(
    facility_variables: list[list[list[Any]]],
    facilities: Iterable[Facility],
    image_data: Iterable[tuple[str, int, str]],
    resolution: int,
    batch: Batch,
    height: int,
):
    mapped_facilities = dict((facility.name, facility) for facility in facilities)
    for sheet in facility_variables:
        headers = sheet[0]
        content = sheet[1:]
        for x in range(1, len(headers)):
            for y in range(len(content)):
                if content[y][0] is None:
                    break
                setattr(mapped_facilities[content[y][0]], headers[x], content[y][x])
    
    named_maps: dict[str, list[list[int]]] = {}
    for image_path, band_to_note, image_name in image_data:
        image = open_image(image_path, 'r')
        pixels = list(image.getdata(band=band_to_note))
        named_maps[image_name] = tuple(
            tuple(pixels[image.width * i:image.width * (i + 1)])
            for i in range(len(pixels) // image.width)
        )
    
    merged_map: list[list[dict[str, int]]] = [
        [defaultdict(int) for _ in row[::resolution]]
        for row in list(named_maps.values())[0][::resolution]
    ]
    for name, map in named_maps.items():
        for x in range(0, len(map[0]), resolution):
            for y in range(0, len(map), resolution):
                merged_map[y // resolution][x // resolution][name] = map[y][x]
        named_maps[name] = tuple(
            tuple(map[y][::resolution])
            for y in range(0, len(map), resolution)
        )
    
    grid = Grid(
        named_maps=named_maps,
        merged_map=merged_map,
        facilities=facilities,
        resolution=resolution,
        height=height,
        batch=batch,
    )

    return grid, named_maps

def get_directions(path, sheet, width, height):
    data = parse_xlsx(path, sheet)[0]
    directions = defaultdict(set)
    for y in range(height):
        for x in range(width):
            directions[data[y][x]].add((x, y))
    return directions
